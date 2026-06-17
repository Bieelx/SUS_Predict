"""XLSX and CSV/GZ export utilities."""
import gzip
import re
from io import BytesIO


def slug_filename(texto: str) -> str:
    texto = re.sub(r"\s+", "_", (texto or "").strip())
    texto = re.sub(r"[^A-Za-z0-9_.-]", "", texto)
    return texto[:80] or "sus_predict"


def csv_gz_bytes(df) -> bytes:
    bio = BytesIO()
    with gzip.GzipFile(fileobj=bio, mode="wb", compresslevel=6) as gz:
        gz.write(df.to_csv(index=False).encode("utf-8"))
    return bio.getvalue()


def xlsx_bytes(resultado: dict) -> bytes:
    from openpyxl import Workbook

    meta  = resultado.get("meta",  {}) or {}
    stats = resultado.get("stats", {}) or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Campo", "Valor"])
    for k in ("sistema","uf","cidade","ibge","ano_ini","ano_fim",
              "doenca_cod","gerado_em","modelo","ano_max_confiavel"):
        if k in meta:
            ws.append([k, meta[k]])
    ws.append([])
    for k in ("total","media_anual","variacao_pct","anos_analisados",
              "prox_previsao","prox_lower","prox_upper"):
        ws.append([k, stats.get(k)])

    ws2 = wb.create_sheet("Serie")
    ws2.append(["ano","total","tipo","lower","upper"])
    for i in (resultado.get("serie_com_previsao") or []):
        ws2.append([i.get("ano"),i.get("total"),i.get("tipo"),i.get("lower"),i.get("upper")])

    ws3 = wb.create_sheet("Sexo")
    ws3.append(["sexo","pct"])
    for i in (resultado.get("distribuicao_sexo") or []):
        ws3.append([i.get("sexo"),i.get("pct")])

    ws4 = wb.create_sheet("FaixaEtaria")
    ws4.append(["faixa","pct"])
    for i in (resultado.get("distribuicao_faixa_etaria") or []):
        ws4.append([i.get("faixa"),i.get("pct")])

    ws5 = wb.create_sheet("TopCausas")
    ws5.append(["causa","pct"])
    for i in (resultado.get("top_causas") or []):
        ws5.append([i.get("causa"),i.get("pct")])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
