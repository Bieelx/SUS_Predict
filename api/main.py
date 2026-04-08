"""
╔══════════════════════════════════════════════════════════════╗
║         SUS Predict — Backend FastAPI                        ║
║         Projeto SUS Predict — FIAP TCC 2025/2026             ║
╚══════════════════════════════════════════════════════════════╝

Inicie com:
    # Use o venv do projeto (Python 3.12, exigido pelo PySUS)
    source ../venv/bin/activate
    pip install -r requirements_api.txt
    uvicorn main:app --reload --port 8000

Capacidades detectadas em runtime:
    PYSUS_OK   → obrigatório: baixa dados reais do FTP do DATASUS (requer Python 3.12 + venv/)
    PROPHET_OK → usa Prophet para previsão (com intervalo de confiança); fallback → OLS
"""

import sys, uuid, time, random, json, shutil, logging, gc
from pathlib import Path
from datetime import datetime
from io import BytesIO
import re
import os
import gzip
import urllib.request
import urllib.error
import urllib.parse

# ── Limites de anos com dados consolidados no DATASUS ─────────────────────────
# O DATASUS publica dados em versão "preliminar" e só consolida após meses/anos.
# SINAN tem a maior defasagem: notificações chegam com 2-3 anos de atraso.
# Usar dados além desses limites pode causar inconsistências graves (ex: mesmo
# ano exibindo valores totalmente diferentes dependendo do intervalo selecionado).
_ANO_ATUAL = datetime.now().year
ANO_MAXIMO_CONFIAVEL: dict = {
    "SINAN":  _ANO_ATUAL - 1,   # ex: 2026 → máximo 2025 (defasagem ~1 ano)
    "SIM":    _ANO_ATUAL - 1,   # ex: 2026 → máximo 2025
    "SIH":    _ANO_ATUAL - 1,
    "SINASC": _ANO_ATUAL - 1,
    "SIA":    _ANO_ATUAL - 1,
}

from fastapi import FastAPI, BackgroundTasks, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("sus_predict")

# ── Adiciona raiz do projeto ao path para importar datasus.py ─────────────────
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

# ── Detecta PySUS ─────────────────────────────────────────────────────────────
try:
    from pysus.online_data.SIH    import download as _sih_dl
    from pysus.online_data.SIM    import download as _sim_dl
    from pysus.online_data.SINASC import download as _sinasc_dl
    from pysus.online_data.SINAN  import download as _sinan_dl
    from datasus import to_df, ESTADOS as _EST, CIDADES as _CID
    import pandas as pd
    PYSUS_OK = True
    log.info("✅  PySUS disponível — modo de dados reais ativo")
except Exception as e:
    PYSUS_OK = False
    log.warning(f"⚠️  PySUS não disponível ({e}) — backend requer Python 3.12 + venv do projeto")

# ── Detecta Prophet ───────────────────────────────────────────────────────────
try:
    from prophet import Prophet
    if not PYSUS_OK:
        import pandas as pd          # garante pandas mesmo sem PySUS
    PROPHET_OK = True
    log.info("✅  Prophet disponível — previsão avançada ativa")
except Exception as e:
    PROPHET_OK = False
    log.warning(f"⚠️  Prophet não disponível ({e}) — usando regressão OLS")

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="SUS Predict API", version="2.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

jobs: dict = {}
TEMP_DIR = Path("./temp_data")
TEMP_DIR.mkdir(exist_ok=True)

_SP_SAO_PAULO_IBGE6 = "355030"

# Coordenadas (fatuais) de municípios. Por enquanto apenas o que precisamos para SP;
# novos municípios podem ser adicionados conforme o projeto crescer.
_IBGE6_COORDS = {
    "355030": {"lat": -23.5505, "lon": -46.6333},  # São Paulo (SP)
}

def _env_int(name: str, default: int) -> int:
    try:
        return int(os.getenv(name, str(default)))
    except Exception:
        return default


def _env_bool(name: str, default: bool) -> bool:
    v = os.getenv(name)
    if v is None:
        return default
    return str(v).strip().lower() in ("1", "true", "yes", "y", "on")


def _supabase_headers(service_key: str) -> dict:
    return {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
    }


def _supabase_request(method: str, url: str, service_key: str, body: bytes | None = None, extra_headers: dict | None = None) -> tuple[int, bytes]:
    headers = _supabase_headers(service_key)
    if extra_headers:
        headers.update(extra_headers)
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return resp.getcode(), resp.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read() if hasattr(e, "read") else b""

def _supabase_get_json(url: str, service_key: str) -> object:
    code, payload = _supabase_request("GET", url, service_key, body=None, extra_headers={"Accept": "application/json"})
    if code >= 300:
        msg = payload.decode("utf-8", errors="ignore")[:500]
        raise RuntimeError(f"Supabase GET falhou (HTTP {code}): {msg}")
    try:
        return json.loads(payload.decode("utf-8"))
    except Exception:
        return None


def _supabase_upsert_rows(supabase_url: str, service_key: str, table: str, rows: list[dict], prefer_merge: bool = True) -> None:
    if not rows:
        return
    url = f"{supabase_url.rstrip('/')}/rest/v1/{table}"
    headers = {"Content-Type": "application/json"}
    if prefer_merge:
        headers["Prefer"] = "resolution=merge-duplicates"
    code, payload = _supabase_request("POST", url, service_key, body=json.dumps(rows).encode("utf-8"), extra_headers=headers)
    if code >= 300:
        msg = payload.decode("utf-8", errors="ignore")[:500]
        raise RuntimeError(f"Supabase upsert falhou em {table} (HTTP {code}): {msg}")


def _supabase_upload_raw(supabase_url: str, service_key: str, bucket: str, object_path: str, gz_bytes: bytes) -> None:
    # PUT com upsert
    url = f"{supabase_url.rstrip('/')}/storage/v1/object/{bucket}/{object_path}"
    headers = {"Content-Type": "application/gzip", "x-upsert": "true"}
    code, payload = _supabase_request("PUT", url, service_key, body=gz_bytes, extra_headers=headers)
    if code >= 300:
        msg = payload.decode("utf-8", errors="ignore")[:500]
        raise RuntimeError(f"Supabase Storage upload falhou (HTTP {code}): {msg}")


def _csv_gz_bytes_from_df(df: "pd.DataFrame") -> bytes:
    bio = BytesIO()
    with gzip.GzipFile(fileobj=bio, mode="wb", compresslevel=6) as gz:
        # CSV é mais compatível no free (sem depender de extensões no Postgres)
        gz.write(df.to_csv(index=False).encode("utf-8"))
    return bio.getvalue()


def _maybe_sync_to_supabase(job_id: str, resultado: dict, df: "pd.DataFrame | None") -> None:
    """
    Sincroniza dados tratados no Postgres e (opcionalmente) um bruto compactado no Storage.
    Tudo é opt-in via env vars.
    """
    supabase_url = os.getenv("SUPABASE_URL", "").strip()
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not supabase_url or not service_key:
        return

    bucket = os.getenv("SUPABASE_BUCKET_RAW", "datasus-raw").strip() or "datasus-raw"
    max_raw_bytes = _env_int("SUPABASE_RAW_MAX_BYTES", 250_000_000)  # ajuste pro seu plano free
    enable_raw = _env_bool("SUPABASE_ENABLE_RAW_UPLOAD", True)

    meta = resultado.get("meta", {}) or {}
    stats = resultado.get("stats", {}) or {}

    run_row = {
        "run_id": job_id,
        "sistema": meta.get("sistema"),
        "uf": meta.get("uf"),
        "cidade": meta.get("cidade"),
        "ibge6": str(meta.get("ibge", "") or "")[:6] or None,
        "ano_ini": meta.get("ano_ini"),
        "ano_fim": meta.get("ano_fim"),
        "doenca_cod": meta.get("doenca_cod") or None,
        "modelo": meta.get("modelo"),
        "gerado_em": meta.get("gerado_em"),
        "raw_bucket": None,
        "raw_object_path": None,
        "raw_bytes": None,
    }

    # ── Upsert do run (metadados) ───────────────────────────────────────────
    _supabase_upsert_rows(supabase_url, service_key, "datasus_runs", [run_row], prefer_merge=True)

    # ── Upsert das tabelas tratadas ─────────────────────────────────────────
    serie_rows = []
    for item in (resultado.get("serie_com_previsao") or []):
        serie_rows.append({
            "run_id": job_id,
            "ano": item.get("ano"),
            "tipo": item.get("tipo"),
            "total": item.get("total"),
            "lower": item.get("lower"),
            "upper": item.get("upper"),
        })
    _supabase_upsert_rows(supabase_url, service_key, "datasus_serie", serie_rows, prefer_merge=True)

    sexo_rows = [{"run_id": job_id, "sexo": i.get("sexo"), "pct": i.get("pct")} for i in (resultado.get("distribuicao_sexo") or [])]
    _supabase_upsert_rows(supabase_url, service_key, "datasus_sexo", sexo_rows, prefer_merge=True)

    faixa_rows = [{"run_id": job_id, "faixa": i.get("faixa"), "pct": i.get("pct")} for i in (resultado.get("distribuicao_faixa_etaria") or [])]
    _supabase_upsert_rows(supabase_url, service_key, "datasus_faixa_etaria", faixa_rows, prefer_merge=True)

    causas_rows = [{"run_id": job_id, "causa": i.get("causa"), "pct": i.get("pct")} for i in (resultado.get("top_causas") or [])]
    _supabase_upsert_rows(supabase_url, service_key, "datasus_top_causas", causas_rows, prefer_merge=True)

    # ── Upload do bruto slim (opcional e com budget) ────────────────────────
    if not enable_raw or df is None or df.empty:
        return

    try:
        # só o mínimo do dataframe (sem colunas extra) + _ano
        sistema = str(meta.get("sistema") or "")
        cols = [c for c in (COLS_MINIMAS.get(sistema, []) + ["_ano"]) if c in df.columns]
        raw_df = df[cols].copy() if cols else df.copy()
        gz_bytes = _csv_gz_bytes_from_df(raw_df)
        if len(gz_bytes) > max_raw_bytes:
            log.info(f"  ℹ️  Bruto compactado com {len(gz_bytes):,} bytes excede budget ({max_raw_bytes:,}). Pulando upload do bruto.")
            return

        object_path = f"raw/{meta.get('sistema')}/{meta.get('uf')}/{_slug_filename(meta.get('cidade',''))}/{meta.get('ano_ini')}-{meta.get('ano_fim')}/{job_id}.csv.gz"
        _supabase_upload_raw(supabase_url, service_key, bucket, object_path, gz_bytes)

        # registra objeto (Storage) no Postgres
        raw_obj_row = {
            "run_id": job_id,
            "ano": None,
            "bucket": bucket,
            "object_path": object_path,
            "bytes": len(gz_bytes),
        }
        _supabase_upsert_rows(supabase_url, service_key, "datasus_raw_objects", [raw_obj_row], prefer_merge=True)

        # atualiza run com referência do objeto
        run_row2 = dict(run_row)
        run_row2.update({"raw_bucket": bucket, "raw_object_path": object_path, "raw_bytes": len(gz_bytes)})
        _supabase_upsert_rows(supabase_url, service_key, "datasus_runs", [run_row2], prefer_merge=True)
        log.info(f"  ✅  Bruto enviado ao Supabase Storage: {bucket}/{object_path} ({len(gz_bytes):,} bytes)")
    except Exception as e:
        log.warning(f"  ⚠️  Falha ao enviar bruto ao Supabase: {e}")


def _try_load_from_supabase(req: dict) -> dict | None:
    """
    Tenta montar o mesmo payload de /api/resultado/{job_id} a partir do Supabase,
    evitando re-download do DATASUS quando já existe um run idêntico.
    """
    supabase_url = os.getenv("SUPABASE_URL", "").strip()
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not supabase_url or not service_key:
        return None

    if not _env_bool("SUPABASE_ENABLE_CACHE_READ", True):
        return None

    sistema = req.get("sistema")
    uf = req.get("uf")
    cidade = req.get("cidade")
    ibge6 = str(req.get("ibge") or "")[:6] or None
    ano_ini = int(req.get("ano_ini"))
    ano_fim = int(req.get("ano_fim"))
    doenca_cod = (req.get("doenca_cod") or "").strip() or None

    # Busca run_id correspondente
    filters = [
        ("sistema", f"eq.{sistema}"),
        ("uf", f"eq.{uf}"),
        ("cidade", f"eq.{cidade}"),
        ("ano_ini", f"eq.{ano_ini}"),
        ("ano_fim", f"eq.{ano_fim}"),
    ]
    if ibge6:
        filters.append(("ibge6", f"eq.{ibge6}"))
    else:
        filters.append(("ibge6", "is.null"))

    if doenca_cod:
        filters.append(("doenca_cod", f"eq.{doenca_cod}"))
    else:
        filters.append(("doenca_cod", "is.null"))

    q = "&".join([f"select=*" ] + [f"{k}={urllib.parse.quote(v, safe='._-')}" for k, v in filters] + ["limit=1"])
    runs_url = f"{supabase_url.rstrip('/')}/rest/v1/datasus_runs?{q}"
    runs = _supabase_get_json(runs_url, service_key)
    if not isinstance(runs, list) or not runs:
        return None

    run = runs[0]
    run_id = run.get("run_id")
    if not run_id:
        return None

    # Busca tabelas tratadas
    def _sel(table: str) -> list:
        url = f"{supabase_url.rstrip('/')}/rest/v1/{table}?select=*&run_id=eq.{urllib.parse.quote(run_id, safe='')}"
        data = _supabase_get_json(url, service_key)
        return data if isinstance(data, list) else []

    serie_rows = _sel("datasus_serie")
    sexo_rows = _sel("datasus_sexo")
    faixa_rows = _sel("datasus_faixa_etaria")
    causas_rows = _sel("datasus_top_causas")

    # Reconstrói stats a partir da série real + previsto
    serie_prev = sorted(
        [
            {
                "ano": r.get("ano"),
                "tipo": r.get("tipo"),
                "total": r.get("total"),
                "lower": r.get("lower"),
                "upper": r.get("upper"),
            }
            for r in serie_rows
        ],
        key=lambda x: (x["ano"] or 0, 0 if x["tipo"] == "real" else 1),
    )
    serie_real = [s for s in serie_prev if s.get("tipo") == "real"]

    total = int(sum(int(s.get("total") or 0) for s in serie_real)) if serie_real else 0
    anos_n = (ano_fim - ano_ini + 1)
    media = int(total // anos_n) if anos_n > 0 else total
    if len(serie_real) >= 2 and (serie_real[0].get("total") or 0) > 0:
        variacao = round(((serie_real[-1]["total"] - serie_real[0]["total"]) / serie_real[0]["total"]) * 100, 1)
    else:
        variacao = 0.0

    prox = next((s.get("total") for s in serie_prev if s.get("tipo") == "previsto"), None)
    prox_low = next((s.get("lower") for s in serie_prev if s.get("tipo") == "previsto"), None)
    prox_up = next((s.get("upper") for s in serie_prev if s.get("tipo") == "previsto"), None)

    resultado = {
        "meta": {
            "sistema": run.get("sistema"),
            "uf": run.get("uf"),
            "cidade": run.get("cidade"),
            "ibge": (run.get("ibge6") or ""),
            "ano_ini": run.get("ano_ini"),
            "ano_fim": run.get("ano_fim"),
            "doenca_cod": run.get("doenca_cod") or "",
            "gerado_em": run.get("gerado_em"),
            "dados_reais": True,
            "dados_completos": True,
            "ano_max_confiavel": run.get("ano_max_confiavel"),
            "modelo": run.get("modelo"),
            "supabase_cache": True,
            "run_id": run_id,
        },
        "stats": {
            "total": total,
            "media_anual": media,
            "variacao_pct": variacao,
            "anos_analisados": anos_n,
            "prox_previsao": prox,
            "prox_lower": prox_low,
            "prox_upper": prox_up,
        },
        "serie_temporal": serie_real,
        "serie_com_previsao": serie_prev,
        "distribuicao_sexo": [{"sexo": r.get("sexo"), "pct": r.get("pct")} for r in sexo_rows],
        "distribuicao_faixa_etaria": [{"faixa": r.get("faixa"), "pct": r.get("pct")} for r in faixa_rows],
        "top_causas": [{"causa": r.get("causa"), "pct": r.get("pct")} for r in causas_rows],
    }
    return resultado


def _list_runs_from_supabase(sistema: str | None = None, limit: int = 200) -> list[dict]:
    supabase_url = os.getenv("SUPABASE_URL", "").strip()
    service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not supabase_url or not service_key:
        return []

    limit = max(1, min(int(limit), 500))
    base = f"{supabase_url.rstrip('/')}/rest/v1/datasus_runs"

    # Seleciona só campos necessários para UI
    select = "run_id,sistema,uf,cidade,ibge6,ano_ini,ano_fim,doenca_cod,gerado_em,modelo"
    query = [("select", select), ("order", "created_at.desc"), ("limit", str(limit))]
    if sistema:
        query.append(("sistema", f"eq.{sistema}"))

    url = f"{base}?{urllib.parse.urlencode(query, safe='.,:_-')}"
    data = _supabase_get_json(url, service_key)
    if not isinstance(data, list):
        return []

    out = []
    for row in data:
        ibge6 = str(row.get("ibge6") or "")[:6]
        coords = _IBGE6_COORDS.get(ibge6)
        out.append({
            "run_id": row.get("run_id"),
            "sistema": row.get("sistema"),
            "uf": row.get("uf"),
            "cidade": row.get("cidade"),
            "ibge6": ibge6 or None,
            "ano_ini": row.get("ano_ini"),
            "ano_fim": row.get("ano_fim"),
            "doenca_cod": row.get("doenca_cod") or "",
            "gerado_em": row.get("gerado_em"),
            "modelo": row.get("modelo"),
            "lat": coords.get("lat") if coords else None,
            "lon": coords.get("lon") if coords else None,
        })
    return out


# ══════════════════════════════════════════════════════════════════════════════
#  DADOS ESTÁTICOS (espelho do datasus.py — usado quando PySUS está indisponível)
# ══════════════════════════════════════════════════════════════════════════════

ESTADOS_FALLBACK = [
    {"sigla":"SP","nome":"São Paulo"},{"sigla":"RJ","nome":"Rio de Janeiro"},
    {"sigla":"MG","nome":"Minas Gerais"},{"sigla":"RS","nome":"Rio Grande do Sul"},
    {"sigla":"PR","nome":"Paraná"},{"sigla":"BA","nome":"Bahia"},
    {"sigla":"CE","nome":"Ceará"},{"sigla":"PE","nome":"Pernambuco"},
    {"sigla":"GO","nome":"Goiás"},{"sigla":"AM","nome":"Amazonas"},
    {"sigla":"SC","nome":"Santa Catarina"},{"sigla":"PA","nome":"Pará"},
    {"sigla":"MA","nome":"Maranhão"},{"sigla":"MT","nome":"Mato Grosso"},
    {"sigla":"MS","nome":"Mato Grosso do Sul"},{"sigla":"PB","nome":"Paraíba"},
    {"sigla":"RN","nome":"Rio Grande do Norte"},{"sigla":"PI","nome":"Piauí"},
    {"sigla":"AL","nome":"Alagoas"},{"sigla":"SE","nome":"Sergipe"},
    {"sigla":"RO","nome":"Rondônia"},{"sigla":"TO","nome":"Tocantins"},
    {"sigla":"AC","nome":"Acre"},{"sigla":"AP","nome":"Amapá"},
    {"sigla":"RR","nome":"Roraima"},{"sigla":"DF","nome":"Distrito Federal"},
    {"sigla":"ES","nome":"Espírito Santo"},
]

CIDADES_FALLBACK = {
    "SP":[("São Paulo","3550308"),("Campinas","3509502"),("Guarulhos","3518800"),("Santo André","3547809"),("Osasco","3534401"),("Ribeirão Preto","3543402"),("Sorocaba","3552205"),("Santos","3548100"),("São José dos Campos","3549904"),("Barueri","3505708")],
    "RJ":[("Rio de Janeiro","3304557"),("Niterói","3303302"),("Nova Iguaçu","3303500"),("Duque de Caxias","3301702"),("São Gonçalo","3304904"),("Petrópolis","3303906"),("Volta Redonda","3306701"),("Campos dos Goytacazes","3301009")],
    "MG":[("Belo Horizonte","3106200"),("Uberlândia","3170206"),("Contagem","3118601"),("Juiz de Fora","3136702"),("Betim","3106705"),("Montes Claros","3143302"),("Uberaba","3170107")],
    "RS":[("Porto Alegre","4314902"),("Caxias do Sul","4305108"),("Pelotas","4314407"),("Canoas","4304606"),("Santa Maria","4316907"),("Novo Hamburgo","4313409")],
    "PR":[("Curitiba","4106902"),("Londrina","4113700"),("Maringá","4115200"),("Ponta Grossa","4119905"),("Cascavel","4104808"),("São José dos Pinhais","4125506")],
    "BA":[("Salvador","2927408"),("Feira de Santana","2910800"),("Vitória da Conquista","2933307"),("Camaçari","2905701"),("Itabuna","2914802")],
    "CE":[("Fortaleza","2304400"),("Caucaia","2303709"),("Juazeiro do Norte","2307304"),("Sobral","2312908"),("Maracanaú","2307650")],
    "PE":[("Recife","2611606"),("Caruaru","2604106"),("Olinda","2609600"),("Petrolina","2611101"),("Paulista","2610707")],
    "GO":[("Goiânia","5208707"),("Aparecida de Goiânia","5201405"),("Anápolis","5201108"),("Rio Verde","5218805")],
    "AM":[("Manaus","1302603"),("Parintins","1303403"),("Itacoatiara","1301902"),("Manacapuru","1302504")],
    "SC":[("Florianópolis","4205407"),("Joinville","4209102"),("Blumenau","4202404"),("Chapecó","4204202"),("Itajaí","4207304")],
    "PA":[("Belém","1501402"),("Ananindeua","1500800"),("Santarém","1506807"),("Marabá","1504208")],
    "MA":[("São Luís","2111300"),("Imperatriz","2105302"),("Timon","2112209")],
    "MT":[("Cuiabá","5103403"),("Várzea Grande","5108402"),("Rondonópolis","5107602"),("Sinop","5107909")],
    "MS":[("Campo Grande","5002704"),("Dourados","5003702"),("Três Lagoas","5008305")],
    "PB":[("João Pessoa","2507507"),("Campina Grande","2504009"),("Santa Rita","2513703")],
    "RN":[("Natal","2408102"),("Mossoró","2408003"),("Parnamirim","2403251")],
    "PI":[("Teresina","2211001"),("Parnaíba","2207702"),("Picos","2208007")],
    "AL":[("Maceió","2704302"),("Arapiraca","2700300"),("Rio Largo","2707701")],
    "SE":[("Aracaju","2800308"),("Nossa Sra. do Socorro","2804805"),("Lagarto","2803500")],
    "RO":[("Porto Velho","1100205"),("Ji-Paraná","1100122"),("Ariquemes","1100023")],
    "TO":[("Palmas","1721000"),("Araguaína","1702109"),("Gurupi","1709500")],
    "AC":[("Rio Branco","1200401"),("Cruzeiro do Sul","1200203")],
    "AP":[("Macapá","1600303"),("Santana","1600600")],
    "RR":[("Boa Vista","1400100"),("Rorainópolis","1400472")],
    "DF":[("Brasília","5300108")],
    "ES":[("Vitória","3205309"),("Vila Velha","3205200"),("Serra","3205010"),("Cariacica","3201308")],
}

# ── Lista local de estados — independente do PySUS ────────────────────────────
_ESTADOS_LOCAL: list[tuple[str, str]] = [
    ("AC", "Acre"),           ("AL", "Alagoas"),      ("AM", "Amazonas"),
    ("AP", "Amapá"),          ("BA", "Bahia"),         ("CE", "Ceará"),
    ("DF", "Distrito Federal"),("ES", "Espírito Santo"),("GO", "Goiás"),
    ("MA", "Maranhão"),       ("MG", "Minas Gerais"),  ("MS", "Mato Grosso do Sul"),
    ("MT", "Mato Grosso"),    ("PA", "Pará"),          ("PB", "Paraíba"),
    ("PE", "Pernambuco"),     ("PI", "Piauí"),         ("PR", "Paraná"),
    ("RJ", "Rio de Janeiro"), ("RN", "Rio Grande do Norte"),("RO", "Rondônia"),
    ("RR", "Roraima"),        ("RS", "Rio Grande do Sul"),  ("SC", "Santa Catarina"),
    ("SE", "Sergipe"),        ("SP", "São Paulo"),     ("TO", "Tocantins"),
]

# ── Cache de municípios por UF — preenchido via API do IBGE em runtime ────────
_MUNICIPIOS_CACHE: dict[str, list[tuple[str, str]]] = {}


def _buscar_municipios_ibge(uf: str) -> list[tuple[str, str]]:
    """
    Retorna todos os municípios de uma UF como [(nome, ibge7)].

    Fonte primária: API IBGE localidades (5.570 municípios, ordenados por nome).
    Fallback:       CIDADES_FALLBACK (lista reduzida hardcoded).

    O resultado é cacheado em memória por sessão de servidor.
    O código IBGE retornado tem 7 dígitos; o DATASUS usa os 6 primeiros.
    """
    uf = uf.upper()
    if uf in _MUNICIPIOS_CACHE:
        return _MUNICIPIOS_CACHE[uf]

    url = (
        f"https://servicodados.ibge.gov.br/api/v1/localidades"
        f"/estados/{uf}/municipios?orderBy=nome"
    )
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "SUSPredict/2.0"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            dados = json.loads(resp.read().decode("utf-8"))
        resultado = [(m["nome"], str(m["id"])) for m in dados]
        _MUNICIPIOS_CACHE[uf] = resultado
        log.info(f"IBGE: {len(resultado)} municípios carregados para {uf}")
        return resultado
    except Exception as exc:
        log.warning(f"IBGE API falhou para {uf} ({exc}) — usando CIDADES_FALLBACK")
        return CIDADES_FALLBACK.get(uf, [])


POPULACAO_REL = {
    "SP":1.00,"MG":0.50,"RJ":0.45,"BA":0.37,"PR":0.30,"RS":0.28,"PE":0.25,
    "CE":0.23,"PA":0.20,"SC":0.18,"MA":0.17,"GO":0.16,"AM":0.15,"DF":0.14,
    "ES":0.10,"PB":0.10,"RN":0.09,"MT":0.09,"MS":0.08,"PI":0.08,"AL":0.08,
    "SE":0.06,"RO":0.05,"TO":0.04,"AC":0.03,"AP":0.025,"RR":0.02,
}

BASE_SISTEMA = {"SIM":130_000,"SIH":900_000,"SINASC":160_000,"SIA":2_500_000}

COL_MUNICIPIO = {"SIM":"CODMUNRES","SIH":"MUNIC_RES","SINASC":"CODMUNNASC","SIA":"CODUFMUN","SINAN":"ID_MUNICIP"}
COL_SEXO      = {"SIM":"SEXO","SIH":"SEXO","SINASC":"SEXO","SIA":"SEXO","SINAN":"CS_SEXO"}

# Colunas mínimas por sistema — só essas são mantidas em RAM após o download.
# Reduz DataFrame de 50-80 colunas para 4-5, economizando até 95% de RAM.
COLS_MINIMAS = {
    "SIM":    ["CODMUNRES", "SEXO", "IDADE",    "CAUSABAS"],
    "SIH":    ["MUNIC_RES", "SEXO", "IDADE",    "DIAG_PRINC"],
    "SINASC": ["CODMUNNASC","SEXO", "IDADEMAE", "PARTO"],
    "SIA":    ["CODUFMUN",  "SEXO", "PA_PROC_ID"],
    "SINAN":  ["ID_MUNICIP","CS_SEXO","CS_FAIXA_ETARIA","CLASSI_FIN"],
}


def _ler_slim(raw, sistema: str, ibge6: str, *, strict_municipio: bool = False) -> "pd.DataFrame | None":
    """
    Substitui to_df() com leitura otimizada de parquet:
    1. Column pruning: lê só as 4-5 colunas necessárias do disco (evita 80+ colunas em RAM)
    2. Filtro de município aplicado após a leitura slim (com fallback para UF inteira)

    Para SINAN DENG 2024 (10M linhas × 80 cols ~4GB) → passa a ser 10M × 5 cols ~200MB,
    depois do filtro de município → ~300K linhas × 5 cols = ~15 MB.
    """
    from pathlib import Path as _Path

    colunas = COLS_MINIMAS.get(sistema, [])
    col_mun = COL_MUNICIPIO.get(sistema)

    def _ler_parquetset(ps) -> "pd.DataFrame | None":
        """Lê um ParquetSet selecionando só as colunas necessárias."""
        base = _Path(ps.path)
        arqs = sorted(base.glob("*.parquet")) if base.is_dir() else [base]
        partes = []
        for arq in arqs:
            df = None
            # Tenta pyarrow primeiro (melhor suporte a column pruning)
            if colunas:
                try:
                    df = pd.read_parquet(str(arq), engine="pyarrow", columns=colunas)
                except Exception:
                    pass
            # Fallback: fastparquet com seleção de colunas
            if df is None and colunas:
                try:
                    df = pd.read_parquet(str(arq), engine="fastparquet", columns=colunas)
                except Exception:
                    pass
            # Último recurso: lê tudo e seleciona depois
            if df is None:
                try:
                    df = pd.read_parquet(str(arq), engine="fastparquet")
                    if colunas:
                        pres = [c for c in colunas if c in df.columns]
                        df = df[pres] if pres else df
                except Exception as exc:
                    log.warning(f"    ⚠️  Falha ao ler {arq.name}: {exc}")
                    continue

            if df is not None and not df.empty:
                partes.append(df)
            del df

        return pd.concat(partes, ignore_index=True) if partes else None

    # ── Normaliza a entrada (ParquetSet único ou lista) ───────────────────────
    if hasattr(raw, "path"):
        df = _ler_parquetset(raw)
    elif isinstance(raw, list):
        partes = []
        for item in raw:
            if hasattr(item, "path"):
                p = _ler_parquetset(item)
                if p is not None and not p.empty:
                    partes.append(p)
                del p
        df = pd.concat(partes, ignore_index=True) if partes else None
    else:
        # Fallback para to_df clássico + slim posterior
        df = to_df(raw)
        if df is not None and colunas:
            pres = [c for c in colunas if c in df.columns]
            if pres:
                df = df[pres]

    if df is None or df.empty:
        return None

    # ── Filtro de município ─────────────────────────────────────────────────
    if ibge6 and col_mun and col_mun in df.columns:
        mask = df[col_mun].astype(str).str[:6] == ibge6[:6]
        if mask.any():
            return df[mask].copy()
        if strict_municipio:
            log.info(f"    📍 Município {ibge6} sem dados na col {col_mun} — retornando vazio (strict)")
            return df.iloc[0:0].copy()
        log.info(f"    📍 Município {ibge6} sem dados na col {col_mun} — usando UF inteira")

    return df


def _slim_municipio(df: "pd.DataFrame", sistema: str, ibge6: str) -> "pd.DataFrame":
    """Utilitário de filtro+slim para DataFrames já em memória (usado como fallback)."""
    if df is None or df.empty:
        return df
    col_mun = COL_MUNICIPIO.get(sistema)
    df_mun  = df
    if ibge6 and col_mun and col_mun in df.columns:
        mask = df[col_mun].astype(str).str[:6] == str(ibge6)[:6]
        if mask.any():
            df_mun = df[mask]
    colunas  = COLS_MINIMAS.get(sistema, [])
    presentes = [c for c in colunas if c in df_mun.columns]
    return df_mun[presentes].copy() if presentes else df_mun


# ══════════════════════════════════════════════════════════════════════════════
#  INVALIDAÇÃO DE CACHE DO PYSUS
# ══════════════════════════════════════════════════════════════════════════════

def _limpar_cache_pysus(sistema: str, anos: list, uf: str = "", doenca_cod: str = "") -> int:
    """
    Remove arquivos parquet em cache do PySUS para os anos solicitados.
    Força re-download da versão mais recente do FTP do DATASUS, evitando
    que dados preliminares/desatualizados sejam exibidos em consultas futuras.

    O PySUS armazena parquets em ~/.pysus/ com nomenclatura baseada nos
    arquivos .dbc originais do DATASUS (ex: DENGBR23.parquet, DOSP22.parquet).
    """
    cache_base = Path.home() / ".pysus"
    if not cache_base.exists():
        log.info("  ℹ️  Cache PySUS (~/.pysus) não encontrado — download direto do FTP")
        return 0

    def _padroes(ano: int) -> list:
        """Gera padrões de nome de arquivo para o sistema/ano/UF."""
        yy = str(ano)[2:]   # ex: 2021 → "21"
        uf_u = uf.upper()
        uf_l = uf.lower()
        if sistema == "SINAN":
            cod = doenca_cod.upper() if doenca_cod else "*"
            return [
                f"*{cod}*{ano}*", f"*{cod}*{yy}*",
                f"*{cod.lower()}*{ano}*", f"*{cod.lower()}*{yy}*",
            ]
        elif sistema == "SIM":
            return [f"*DO{uf_u}{yy}*", f"*DO{uf_l}{yy}*", f"*sim*{yy}*"]
        elif sistema == "SIH":
            return [f"*RD{uf_u}{yy}*", f"*RD{uf_l}{yy}*", f"*sih*{yy}*"]
        elif sistema == "SINASC":
            return [f"*DN{uf_u}{yy}*", f"*DN{uf_l}{yy}*", f"*sinasc*{yy}*"]
        elif sistema == "SIA":
            return [f"*PA{uf_u}{yy}*", f"*PA{uf_l}{yy}*", f"*sia*{yy}*"]
        return []

    removidos = 0
    for ano in anos:
        for padrao in _padroes(ano):
            for arq in cache_base.rglob(padrao):
                try:
                    if arq.is_file():
                        arq.unlink()
                        removidos += 1
                        log.info(f"  🗑️  Cache removido: {arq.name}")
                except Exception as e:
                    log.warning(f"  ⚠️  Não foi possível remover cache {arq.name}: {e}")

    if removidos:
        log.info(f"  ✅  {removidos} arquivo(s) de cache removido(s) — dados frescos garantidos")
    else:
        log.info(f"  ℹ️  Sem cache local para {sistema} — FTP será consultado diretamente")
    return removidos


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS DE EXTRAÇÃO DE DADOS REAIS
# ══════════════════════════════════════════════════════════════════════════════

def _baixar_ano(sistema: str, uf: str, ano: int, doenca_cod: str = None, ibge6: str = ""):
    """
    Baixa 1 ano de dados via PySUS.
    Usa _ler_slim() para fazer column pruning direto no parquet,
    eliminando o gargalo de RAM que trava o Mac com 8 GB.
    """
    try:
        if sistema == "SIM":
            raw = _sim_dl("CID10", uf, ano)
            df  = _ler_slim(raw, "SIM", ibge6)

        elif sistema == "SIH":
            # SIH tem 12 arquivos mensais — aplicamos _ler_slim em cada mês
            # para nunca ter o estado inteiro em RAM ao mesmo tempo.
            frames_mes = []
            for mes in range(1, 13):
                try:
                    raw_mes = _sih_dl(uf, ano, mes, "RD")
                    f = _ler_slim(raw_mes, "SIH", ibge6)
                    if f is not None and not f.empty:
                        frames_mes.append(f)
                    del raw_mes, f
                except Exception:
                    pass
            df = pd.concat(frames_mes, ignore_index=True) if frames_mes else None
            del frames_mes
            if df is not None and not df.empty:
                df["_ano"] = ano
                return df
            return None

        elif sistema == "SINASC":
            raw = _sinasc_dl("DN", uf, ano)
            df  = _ler_slim(raw, "SINASC", ibge6)

        elif sistema == "SINAN":
            return _baixar_sinan(doenca_cod, ano, ibge6=ibge6)

        else:
            return None

        if df is not None and not df.empty:
            df["_ano"] = ano
            return df

    except Exception as e:
        log.warning(f"  ⚠️  {sistema} {uf} {ano}: {e}")
    return None


def _baixar_sinan(doenca_cod: str, ano: int, ibge6: str = ""):
    """
    Baixa dados nacionais do SINAN para uma doença.
    _ler_slim() aplica column pruning + filtro de município no parquet,
    crucial para anos como DENG 2024 com 10M+ linhas nacionais.
    """
    if not doenca_cod:
        return None
    try:
        raw = _sinan_dl(doenca_cod, ano)
        # SINAN é nacional: se não houver match do município, não fazemos fallback para "UF inteira".
        df  = _ler_slim(raw, "SINAN", ibge6, strict_municipio=bool(ibge6))
        if df is not None and not df.empty:
            df["_ano"] = ano
            return df
    except Exception as e:
        log.warning(f"  ⚠️  SINAN {doenca_cod} {ano}: {e}")
    return None


def _filtrar_municipio(df: "pd.DataFrame", sistema: str, ibge6: str) -> "pd.DataFrame":
    """Filtra por município se ibge6 informado."""
    if not ibge6 or df is None or df.empty:
        return df
    col = COL_MUNICIPIO.get(sistema)
    if col and col in df.columns:
        return df[df[col].astype(str).str[:6] == ibge6[:6]]
    return df


def _serie_de_df(df: "pd.DataFrame", ano_ini: int, ano_fim: int) -> list:
    """Conta registros por ano a partir de DataFrame real."""
    serie = []
    for ano in range(ano_ini, ano_fim + 1):
        total = int((df["_ano"] == ano).sum()) if "_ano" in df.columns else 0
        serie.append({"ano": ano, "total": total, "tipo": "real"})
    return serie


def _sexo_de_df(df: "pd.DataFrame", sistema: str) -> list | None:
    """Extrai distribuição de sexo de DataFrame real."""
    col = COL_SEXO.get(sistema, "SEXO")
    if df is None or df.empty or col not in df.columns:
        return None
    # Normaliza: 1/M → Masculino, 2/F → Feminino, ignora Ignorado/inválido
    mapa = {"1":"Masculino","M":"Masculino","m":"Masculino",
            "2":"Feminino", "F":"Feminino", "f":"Feminino"}
    s = df[col].astype(str).str.strip().map(mapa).dropna()
    if len(s) == 0:
        return None
    counts = s.value_counts()
    total = counts.sum()
    return [{"sexo": k, "pct": round(v / total * 100)} for k, v in counts.items()]


def _faixa_de_df(df: "pd.DataFrame", sistema: str) -> list | None:
    """Extrai distribuição por faixa etária de DataFrame real."""
    # SINAN usa CS_FAIXA_ETARIA com código 1-12 (padrão epidemiológico)
    if sistema == "SINAN":
        return _faixa_sinan_de_df(df)

    col = "IDADEMAE" if sistema == "SINASC" else "IDADE"
    if df is None or df.empty or col not in df.columns:
        return None
    try:
        idades = pd.to_numeric(df[col], errors="coerce").dropna()
        if len(idades) == 0:
            return None

        if sistema == "SINASC":
            bins   = [0, 19, 24, 29, 34, 39, 120]
            labels = ["Mães <20","Mães 20–24","Mães 25–29","Mães 30–34","Mães 35–39","Mães 40+"]
        else:
            bins   = [0, 14, 29, 44, 59, 74, 200]
            labels = ["0–14","15–29","30–44","45–59","60–74","75+"]

        # SIM/SIH codificam idade de forma especial (ex: 301 = 1 ano, 410 = 10 anos)
        # Aqui usamos uma heurística simples: se a maioria > 200, decodifica
        if sistema in ("SIM", "SIH") and idades.median() > 200:
            def decodificar(v):
                v = int(v)
                if v < 100:   return v          # dias (neonatal) → 0
                if v < 200:   return 0          # meses → 0
                if v < 300:   return 0          # horas → 0
                return v - 400 if v >= 400 else v - 300
            idades = idades.map(decodificar).clip(0, 120)

        cats   = pd.cut(idades, bins=bins, labels=labels, right=True)
        counts = cats.value_counts().reindex(labels, fill_value=0)
        total  = counts.sum()
        if total == 0:
            return None
        return [{"faixa": f, "pct": round(c / total * 100)} for f, c in counts.items()]
    except Exception as e:
        log.warning(f"  ⚠️  faixa etária: {e}")
        return None


# ── Faixa etária SINAN ────────────────────────────────────────────────────────

# CS_FAIXA_ETARIA: código padrão SINAN → grupo de exibição
_SINAN_FAIXA = {
    "1": "< 1 ano", "2": "1–4", "3": "5–9",
    "4": "10–14",  "5": "15–19",
    "6": "20–29",  "7": "30–39",
    "8": "40–49",  "9": "50–59",
    "10": "60–69", "11": "70–79", "12": "80+",
}

def _faixa_sinan_de_df(df: "pd.DataFrame") -> list | None:
    """Faixa etária para dados SINAN (coluna CS_FAIXA_ETARIA)."""
    col = "CS_FAIXA_ETARIA"
    if df is None or df.empty or col not in df.columns:
        return None
    try:
        faixas = df[col].astype(str).str.strip().map(_SINAN_FAIXA).dropna()
        if len(faixas) == 0:
            return None
        labels = list(dict.fromkeys(_SINAN_FAIXA.values()))   # mantém ordem
        counts = faixas.value_counts().reindex(labels, fill_value=0)
        total  = counts.sum()
        if total == 0:
            return None
        return [{"faixa": f, "pct": round(c / total * 100)} for f, c in counts.items() if c > 0]
    except Exception as e:
        log.warning(f"  ⚠️  faixa_sinan: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  DADOS SINTÉTICOS (mantidos mas NÃO usados em produção — apenas referência)
# ══════════════════════════════════════════════════════════════════════════════

def _serie_sintetica(sistema: str, uf: str, ano_ini: int, ano_fim: int) -> list:
    base   = BASE_SISTEMA.get(sistema, 100_000) // 10
    fator  = POPULACAO_REL.get(uf, 0.05)
    trend  = {"SIM":0.018,"SIH":0.015,"SINASC":-0.012,"SIA":0.025}.get(sistema, 0.015)
    rng    = random.Random(abs(hash(f"{sistema}{uf}")) % 99999)
    return [
        {"ano": ano, "total": int(base * fator * (1 + i * trend) * rng.uniform(0.93, 1.07)), "tipo": "real"}
        for i, ano in enumerate(range(ano_ini, ano_fim + 1))
    ]

def _sexo_sintetico(sistema: str) -> list:
    return {
        "SIM":    [{"sexo":"Masculino","pct":56},{"sexo":"Feminino","pct":44}],
        "SINASC": [{"sexo":"Masculino","pct":51},{"sexo":"Feminino","pct":49}],
        "SIH":    [{"sexo":"Feminino", "pct":53},{"sexo":"Masculino","pct":47}],
        "SIA":    [{"sexo":"Feminino", "pct":57},{"sexo":"Masculino","pct":43}],
    }.get(sistema, [{"sexo":"Masculino","pct":50},{"sexo":"Feminino","pct":50}])

def _faixa_sintetica(sistema: str) -> list:
    return {
        "SIM":    [("0–14",4),("15–29",7),("30–44",11),("45–59",21),("60–74",31),("75+",26)],
        "SIH":    [("0–14",18),("15–29",13),("30–44",15),("45–59",20),("60–74",21),("75+",13)],
        "SINASC": [("Mães <20",16),("Mães 20–24",24),("Mães 25–29",25),("Mães 30–34",20),("Mães 35–39",11),("Mães 40+",4)],
        "SIA":    [("0–14",14),("15–29",18),("30–44",20),("45–59",22),("60–74",17),("75+",9)],
    }.get(sistema, [("0–14",15),("15–29",20),("30–44",22),("45–59",22),("60–74",13),("75+",8)])

def _faixa_sintetica_list(sistema: str) -> list:
    return [{"faixa": f, "pct": p} for f, p in _faixa_sintetica(sistema)]

# ── Mapeamentos para agrupamento de causas ────────────────────────────────────

# Capítulos CID-10 (primeira letra do código)
CID10_CAP = {
    "A": "Infecciosas e parasitárias",
    "B": "Infecciosas e parasitárias",
    "C": "Neoplasias malignas",
    "D": "Neoplasias / sangue",
    "E": "Doenças endócrinas e metabólicas",
    "F": "Transtornos mentais",
    "G": "Sistema nervoso",
    "H": "Olhos / ouvidos",
    "I": "Doenças cardiovasculares",
    "J": "Doenças respiratórias",
    "K": "Doenças digestivas",
    "L": "Doenças da pele",
    "M": "Sistema osteomuscular",
    "N": "Doenças geniturinárias",
    "O": "Gravidez, parto e puerpério",
    "P": "Afecções perinatais",
    "Q": "Malformações congênitas",
    "R": "Sintomas e sinais",
    "S": "Traumatismos",
    "T": "Envenenamentos",
    "V": "Acidentes de trânsito",
    "W": "Causas externas — acidentes",
    "X": "Causas externas — violência",
    "Y": "Causas externas — indeterminadas",
    "Z": "Fatores de saúde",
}

# Grupos de procedimento SIA (2 primeiros dígitos do PA_PROC_ID)
SIA_GRUPO = {
    "01": "Promoção e prevenção",
    "02": "Diagnóstico",
    "03": "Procedimentos clínicos",
    "04": "Procedimentos cirúrgicos",
    "05": "Transplantes",
    "06": "Medicamentos",
    "07": "Órteses e próteses",
    "08": "Ações complementares",
}


def _causas_de_df(df: "pd.DataFrame", sistema: str) -> list | None:
    """
    Extrai distribuição real de causas / diagnósticos do DataFrame do PySUS.
    - SIM    → agrupa CAUSABAS por capítulo CID-10
    - SIH    → agrupa DIAG_PRINC por capítulo CID-10
    - SINASC → distribuição do tipo de parto (PARTO)
    - SIA    → agrupa PA_PROC_ID pelos 2 primeiros dígitos
    """
    try:
        if sistema == "SIM":
            col = "CAUSABAS"
            if col not in df.columns:
                return None
            caps = df[col].astype(str).str[0].str.upper()
            series = caps.map(lambda c: CID10_CAP.get(c, "Outras causas"))

        elif sistema == "SIH":
            col = "DIAG_PRINC"
            if col not in df.columns:
                return None
            caps = df[col].astype(str).str[0].str.upper()
            series = caps.map(lambda c: CID10_CAP.get(c, "Outras causas"))

        elif sistema == "SINASC":
            col = "PARTO"
            if col not in df.columns:
                return None
            mapa = {"1": "Vaginal", "2": "Cesáreo"}
            series = df[col].astype(str).str.strip().map(mapa).dropna()

        elif sistema == "SIA":
            col = "PA_PROC_ID"
            if col not in df.columns:
                return None
            grupos = df[col].astype(str).str.zfill(10).str[:2]
            series = grupos.map(lambda g: SIA_GRUPO.get(g, f"Grupo {g}"))

        elif sistema == "SINAN":
            # Para SINAN o dataset já é de uma doença específica — usamos CLASSI_FIN (classificação
            # final) se disponível, para mostrar distribuição de formas clínicas / desfechos.
            for col_candidata in ("CLASSI_FIN", "CLASSI_FIN_N", "FORMA", "CLASSI_FINL"):
                if col_candidata in df.columns:
                    series = df[col_candidata].astype(str).str.strip()
                    series = series[series.notna() & (series != "nan") & (series != "")]
                    if len(series) > 0:
                        break
            else:
                return None

        else:
            return None

        counts = series.value_counts().head(7)
        total  = counts.sum()
        if total == 0:
            return None

        return [{"causa": k, "pct": round(v / total * 100, 1)} for k, v in counts.items()]

    except Exception as e:
        log.warning(f"  ⚠️  causas_de_df: {e}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
#  MODELO PREDITIVO
# ══════════════════════════════════════════════════════════════════════════════

def _predicao_prophet(serie: list, anos_previsao: int = 3) -> tuple[list, str]:
    """
    Previsão com Meta Prophet.
    Retorna (serie_com_previsao, modelo_usado).
    Para dados anuais: desabilita sazonalidades periódicas, usa só tendência.

    Parâmetros calibrados para séries curtas (3-10 pontos) típicas de dados municipais:
    - changepoint_prior_scale=0.3 → permite capturar tendências reais (0.05 era conservador demais)
    - uncertainty_samples=100     → intervalo de confiança com custo computacional reduzido
    """
    import math
    # log1p estabiliza variância em contagens e melhora ajuste em séries curtas.
    df = pd.DataFrame({
        "ds": pd.to_datetime([f"{s['ano']}-01-01" for s in serie]),
        "y":  [float(math.log1p(max(0, s["total"]))) for s in serie],
    })

    model = Prophet(
        yearly_seasonality      = False,  # dados anuais — sem sazonalidade intra-ano
        weekly_seasonality      = False,
        daily_seasonality       = False,
        uncertainty_samples     = 100,    # era 500 — reduz RAM e tempo de ajuste
        changepoint_prior_scale = 0.3,    # era 0.05 — aumenta flexibilidade da tendência
    )
    model.fit(df)

    try:
        future = model.make_future_dataframe(periods=anos_previsao, freq="YE")
    except Exception:
        future = model.make_future_dataframe(periods=anos_previsao, freq="A")   # pandas < 2.2

    forecast = model.predict(future)

    resultado = []
    for _, row in forecast.iterrows():
        ano  = row["ds"].year
        tipo = "real" if ano <= serie[-1]["ano"] else "previsto"
        yhat = float(row["yhat"])
        ylo  = float(row["yhat_lower"])
        yhi  = float(row["yhat_upper"])
        # volta do log1p para contagens reais
        total = max(0, int(math.expm1(yhat)))
        lower = max(0, int(math.expm1(ylo)))
        upper = max(0, int(math.expm1(yhi)))
        resultado.append({
            "ano":    ano,
            "total":  total,
            "lower":  lower,
            "upper":  upper,
            "tipo":   tipo,
        })
    return resultado, "Prophet(log1p)"


def _predicao_ols(serie: list, anos_previsao: int = 3) -> tuple[list, str]:
    """
    Regressão linear OLS simples (fallback sem numpy/sklearn).
    Calcula IC aproximado via desvio padrão dos resíduos (±1.28σ ≈ 80% bilateral).
    """
    if len(serie) < 2:
        return [dict(s, lower=None, upper=None) for s in serie], "OLS"

    # Ajuste no log1p para melhorar estabilidade em contagens.
    x  = [s["ano"] for s in serie]
    y  = [float(__import__("math").log1p(max(0, s["total"]))) for s in serie]
    n  = len(x)
    xm, ym = sum(x) / n, sum(y) / n
    num = sum((xi - xm) * (yi - ym) for xi, yi in zip(x, y))
    den = sum((xi - xm) ** 2 for xi in x)
    slope     = num / den if den else 0
    intercept = ym - slope * xm

    # Desvio padrão dos resíduos (estimativa do erro de previsão)
    residuos   = [yi - (intercept + slope * xi) for xi, yi in zip(x, y)]
    var_res    = sum(r ** 2 for r in residuos) / max(n - 2, 1)
    sigma      = var_res ** 0.5
    margem     = 1.28 * sigma   # IC ~80% bilateral (aprox. normal)

    # Série real não tem IC aqui (para não confundir com IC de previsão).
    resultado = [dict(s, lower=None, upper=None) for s in serie]
    for i in range(1, anos_previsao + 1):
        ano_p   = max(x) + i
        yhat    = intercept + slope * ano_p
        # IC cresce com distância (no log1p) e depois converte para contagem.
        m = margem * (1 + i * 0.15)
        total_p = max(0, int(__import__("math").expm1(yhat)))
        low_p   = max(0, int(__import__("math").expm1(yhat - m)))
        up_p    = max(0, int(__import__("math").expm1(yhat + m)))
        resultado.append({
            "ano":   ano_p,
            "total": total_p,
            "lower": low_p,
            "upper": up_p,
            "tipo":  "previsto",
        })
    return resultado, "OLS(log1p)"


def _smape(y_true: float, y_pred: float) -> float:
    den = (abs(y_true) + abs(y_pred))
    if den == 0:
        return 0.0
    return 2.0 * abs(y_pred - y_true) / den


def _predicao_holt(serie: list, anos_previsao: int = 3) -> tuple[list, str]:
    """
    Holt (double exponential smoothing) no log1p:
    costuma ser mais estável que Prophet/OLS quando há poucos pontos e tendência suave.
    """
    import math

    if len(serie) < 3:
        return _predicao_ols(serie, anos_previsao)

    y = [math.log1p(max(0, s["total"])) for s in serie]

    # Grid pequeno de hiperparâmetros para evitar custo alto.
    grid = [0.2, 0.35, 0.5, 0.65, 0.8]
    best = None

    def fit(alpha: float, beta: float):
        level = y[0]
        trend = y[1] - y[0]
        errors = []
        for t in range(1, len(y)):
            pred = level + trend
            errors.append(y[t] - pred)
            new_level = alpha * y[t] + (1 - alpha) * (level + trend)
            new_trend = beta * (new_level - level) + (1 - beta) * trend
            level, trend = new_level, new_trend
        # sigma no espaço log para intervalo simples
        if len(errors) >= 2:
            var = sum(e * e for e in errors) / max(len(errors) - 1, 1)
        else:
            var = 0.0
        return level, trend, math.sqrt(var)

    # validação simples: 1-step-ahead nos últimos 2 pontos
    for alpha in grid:
        for beta in grid:
            level, trend, sigma = fit(alpha, beta)
            # Re-fit completo e avalia erro 1 passo (aprox): usa sigma como proxy (mais baixo melhor)
            score = sigma
            if best is None or score < best["score"]:
                best = {"alpha": alpha, "beta": beta, "score": score, "level": level, "trend": trend, "sigma": sigma}

    level = best["level"]
    trend = best["trend"]
    sigma = best["sigma"]

    resultado = [dict(s, lower=None, upper=None) for s in serie]
    # IC ~80% no espaço log
    margem = 1.28 * sigma
    for i in range(1, anos_previsao + 1):
        yhat = level + i * trend
        m = margem * (1 + i * 0.12)
        total = max(0, int(math.expm1(yhat)))
        low = max(0, int(math.expm1(yhat - m)))
        up = max(0, int(math.expm1(yhat + m)))
        resultado.append({"ano": serie[-1]["ano"] + i, "total": total, "lower": low, "upper": up, "tipo": "previsto"})

    return resultado, "Holt(log1p)"


def gerar_predicao(serie: list, anos_previsao: int = 3) -> tuple[list, str]:
    """
    Escolhe Prophet ou OLS conforme disponibilidade e tamanho da série.
    Prophet requer ≥ 4 pontos com valores não-zero para produzir tendência confiável.
    """
    pontos_validos = [s for s in serie if s["total"] > 0]
    # Com poucas observações, Holt tende a ser mais estável.
    if len(pontos_validos) >= 4:
        try:
            return _predicao_holt(serie, anos_previsao)
        except Exception as e:
            log.warning(f"  ⚠️  Holt falhou ({e}) — tentando Prophet/OLS")

    if PROPHET_OK and len(pontos_validos) >= 5:
        try:
            return _predicao_prophet(serie, anos_previsao)
        except Exception as e:
            log.warning(f"  ⚠️  Prophet falhou ({e}) — fallback para OLS")
    if len(pontos_validos) < 4:
        log.info(f"  ℹ️  Série com {len(pontos_validos)} pontos válidos — usando OLS")
    return _predicao_ols(serie, anos_previsao)


def _slug_filename(texto: str) -> str:
    texto = (texto or "").strip()
    texto = re.sub(r"\s+", "_", texto)
    texto = re.sub(r"[^A-Za-z0-9_.-]", "", texto)
    return texto[:80] if texto else "sus_predict"


def _xlsx_bytes(resultado: dict) -> bytes:
    """
    Gera um XLSX em memória com abas:
    - Resumo
    - Serie (real + previsto)
    - Sexo
    - FaixaEtaria
    - TopCausas
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"

    meta = resultado.get("meta", {}) or {}
    stats = resultado.get("stats", {}) or {}

    ws.append(["Campo", "Valor"])
    for k in ("sistema", "uf", "cidade", "ibge", "ano_ini", "ano_fim", "doenca_cod", "gerado_em", "modelo", "ano_max_confiavel"):
        if k in meta:
            ws.append([k, meta.get(k)])
    ws.append([])
    for k in ("total", "media_anual", "variacao_pct", "anos_analisados", "prox_previsao", "prox_lower", "prox_upper"):
        ws.append([k, stats.get(k)])

    ws2 = wb.create_sheet("Serie")
    ws2.append(["ano", "total", "tipo", "lower", "upper"])
    for item in (resultado.get("serie_com_previsao") or []):
        ws2.append([item.get("ano"), item.get("total"), item.get("tipo"), item.get("lower"), item.get("upper")])

    ws3 = wb.create_sheet("Sexo")
    ws3.append(["sexo", "pct"])
    for item in (resultado.get("distribuicao_sexo") or []):
        ws3.append([item.get("sexo"), item.get("pct")])

    ws4 = wb.create_sheet("FaixaEtaria")
    ws4.append(["faixa", "pct"])
    for item in (resultado.get("distribuicao_faixa_etaria") or []):
        ws4.append([item.get("faixa"), item.get("pct")])

    ws5 = wb.create_sheet("TopCausas")
    ws5.append(["causa", "pct"])
    for item in (resultado.get("top_causas") or []):
        ws5.append([item.get("causa"), item.get("pct")])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  JOB PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def _atualizar(job_id: str, pct: int, msg: str):
    jobs[job_id]["progresso"] = pct
    jobs[job_id]["mensagem"]  = msg
    log.info(f"  [{pct:>3}%] {msg}")


def processar_download(job_id: str, req: dict):
    sistema    = req["sistema"]
    uf         = req["uf"]
    ibge       = req["ibge"]
    ano_ini    = req["ano_ini"]
    ano_fim    = req["ano_fim"]
    doenca_cod = req.get("doenca_cod", "")
    anos       = list(range(ano_ini, ano_fim + 1))

    try:
        jobs[job_id]["status"] = "running"
        _atualizar(job_id, 3, "Iniciando pipeline...")

        # ── Download real via PySUS ───────────────────────────────────────────
        if not PYSUS_OK:
            raise RuntimeError(
                "PySUS não disponível. Inicie o backend com o venv/ do projeto (Python 3.12). "
                "Execute: source venv/bin/activate && uvicorn api.main:app --port 8000"
            )

        # ── Validação: bloqueia anos com dados preliminares/incompletos ───────
        ano_max_confiavel = ANO_MAXIMO_CONFIAVEL.get(sistema, _ANO_ATUAL - 2)
        anos_incompletos  = [a for a in anos if a > ano_max_confiavel]
        if anos_incompletos:
            raise ValueError(
                f"Os anos {', '.join(map(str, anos_incompletos))} podem conter dados "
                f"preliminares ou incompletos no DATASUS para o sistema {sistema}. "
                f"O DATASUS consolida os dados de {sistema} com até "
                f"{_ANO_ATUAL - ano_max_confiavel} ano(s) de defasagem. "
                f"Selecione apenas anos até {ano_max_confiavel} para garantir dados completos e confiáveis."
            )

        # ── Invalida cache PySUS (garante dados frescos do FTP) ───────────────
        _atualizar(job_id, 5, "Verificando e limpando cache local do PySUS...")
        _limpar_cache_pysus(sistema, anos, uf=uf, doenca_cod=doenca_cod)

        _atualizar(job_id, 8, "Conectando ao FTP do DATASUS...")
        frames     = []
        total_anos = len(anos)
        ibge6      = str(ibge)[:6] if ibge else ""

        for i, ano in enumerate(anos):
            pct = 10 + int(i / total_anos * 50)   # 10 → 60%
            label = f"SINAN/{doenca_cod.upper()}" if sistema == "SINAN" else f"{sistema} {uf}"
            _atualizar(job_id, pct, f"Baixando {label} — {ano}...")
            # ibge6 é passado para filtrar + slim dentro de _baixar_ano
            df_ano = _baixar_ano(sistema, uf, ano, doenca_cod=doenca_cod, ibge6=ibge6)
            if df_ano is not None:
                frames.append(df_ano)
                log.info(f"    ✔  {ano}: {len(df_ano):,} registros (município/UF já filtrado)")
            del df_ano   # permite GC antes do próximo ano
            gc.collect()

        if not frames:
            raise ValueError(
                f"Nenhum dado encontrado no DATASUS para {sistema} "
                f"{'(' + doenca_cod + ') ' if doenca_cod else ''}"
                f"no período {ano_ini}–{ano_fim}. "
                "Verifique se os dados estão disponíveis e tente um período diferente."
            )

        import pandas as pd
        df = pd.concat(frames, ignore_index=True)
        del frames   # libera lista de frames intermediários
        gc.collect()
        _atualizar(job_id, 62, f"{len(df):,} registros carregados. Processando...")

        # ── Série temporal ────────────────────────────────────────────────────
        _atualizar(job_id, 65, "Calculando série temporal...")
        serie = _serie_de_df(df, ano_ini, ano_fim)

        log.info(f"  📊 Série temporal: {[s['total'] for s in serie]}")
        if all(s["total"] == 0 for s in serie):
            raise ValueError(
                f"Município/UF selecionado não gerou dados para {sistema} "
                f"no período {ano_ini}–{ano_fim}. "
                "Tente ampliar o período ou escolher outro município."
            )

        # ── Modelo preditivo ──────────────────────────────────────────────────
        _atualizar(job_id, 72, "Ajustando modelo preditivo (Prophet)..." if PROPHET_OK else "Calculando regressão OLS...")
        serie_prev, modelo_usado = gerar_predicao(serie, anos_previsao=3)
        _atualizar(job_id, 85, f"Previsão gerada com {modelo_usado}.")

        # ── Distribuições demográficas e causas ───────────────────────────────
        _atualizar(job_id, 88, "Calculando distribuições...")
        sexo   = _sexo_de_df(df, sistema)
        faixas = _faixa_de_df(df, sistema)
        causas = _causas_de_df(df, sistema) or []

        # ── Stats ─────────────────────────────────────────────────────────────
        total    = sum(s["total"] for s in serie)
        anos_n   = ano_fim - ano_ini + 1
        media    = total // anos_n if anos_n > 0 else total
        variacao = round(
            ((serie[-1]["total"] - serie[0]["total"]) / serie[0]["total"]) * 100, 1
        ) if len(serie) >= 2 and serie[0]["total"] > 0 else 0.0
        prox     = next((s["total"] for s in serie_prev if s["tipo"] == "previsto"), None)
        prox_low = next((s.get("lower") for s in serie_prev if s["tipo"] == "previsto"), None)
        prox_up  = next((s.get("upper") for s in serie_prev if s["tipo"] == "previsto"), None)

        # ── Monta resultado final ─────────────────────────────────────────────
        resultado = {
            "meta": {
                "sistema":    sistema,
                "uf":         uf,
                "cidade":     req["cidade"],
                "ibge":       ibge,
                "ano_ini":    ano_ini,
                "ano_fim":    ano_fim,
                "doenca_cod": doenca_cod,
                "gerado_em":  datetime.now().isoformat(),
                "dados_reais":        True,
                "dados_completos":    True,   # só chega aqui se passou a validação de ano
                "ano_max_confiavel":  ano_max_confiavel,
                "modelo":             modelo_usado,
            },
            "stats": {
                "total":           total,
                "media_anual":     media,
                "variacao_pct":    variacao,
                "anos_analisados": anos_n,
                "prox_previsao":   prox,
                "prox_lower":      prox_low,
                "prox_upper":      prox_up,
            },
            "serie_temporal":            serie,
            "serie_com_previsao":        serie_prev,
            "distribuicao_sexo":         sexo,
            "distribuicao_faixa_etaria": faixas,
            "top_causas":                causas,
        }

        # Persiste localmente
        pasta = TEMP_DIR / job_id
        pasta.mkdir(exist_ok=True)
        (pasta / "resultado.json").write_text(
            json.dumps(resultado, ensure_ascii=False, indent=2), encoding="utf-8"
        )

        # ── Sync opcional para Supabase ──────────────────────────────────────
        _atualizar(job_id, 96, "Sincronizando com Supabase..." if os.getenv("SUPABASE_URL") else "Finalizando...")
        try:
            _maybe_sync_to_supabase(job_id, resultado, df)
        except Exception as e:
            log.warning(f"  ⚠️  Sync Supabase falhou (continua local): {e}")

        _atualizar(job_id, 100, "Análise concluída! ✅")
        jobs[job_id]["status"]    = "done"
        jobs[job_id]["resultado"] = resultado
        jobs[job_id]["pasta"]     = str(pasta)

    except Exception as e:
        log.error(f"  ❌  Job {job_id} falhou: {e}", exc_info=True)
        jobs[job_id]["status"]   = "error"
        jobs[job_id]["mensagem"] = str(e)


# ══════════════════════════════════════════════════════════════════════════════
#  MODELS
# ══════════════════════════════════════════════════════════════════════════════

class DownloadRequest(BaseModel):
    sistema:    str
    uf:         str
    cidade:     str
    ibge:       str
    ano_ini:    int
    ano_fim:    int
    doenca_cod: str = ""   # obrigatório apenas quando sistema == "SINAN"
    usar_cache: bool = True


# ══════════════════════════════════════════════════════════════════════════════
#  ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/")
def root():
    return {
        "status":      "ok",
        "app":         "SUS Predict API",
        "version":     "2.0.0",
        "pysus_ok":    PYSUS_OK,
        "prophet_ok":  PROPHET_OK,
        "docs":        "/docs",
    }


@app.get("/api/sistemas")
def get_sistemas():
    return [
        {"codigo":"SIM",    "nome":"SIM — Mortalidade",          "descricao":"Óbitos com causa básica (CID-10)",               "icone":"💀"},
        {"codigo":"SIH",    "nome":"SIH — Internações",          "descricao":"Internações hospitalares financiadas pelo SUS",   "icone":"🏥"},
        {"codigo":"SINASC", "nome":"SINASC — Nascimentos",       "descricao":"Registros de nascidos vivos por município",       "icone":"👶"},
        {"codigo":"SIA",    "nome":"SIA — Ambulatorial",         "descricao":"Produção ambulatorial do SUS",                   "icone":"🩺"},
        {"codigo":"SINAN",  "nome":"SINAN — Doenças Notificáveis","descricao":"Dengue, tuberculose, meningite e +27 agravos",   "icone":"🦠"},
    ]


@app.get("/api/doencas")
def get_doencas():
    """Lista as doenças disponíveis no SINAN (para o seletor do frontend)."""
    if not PYSUS_OK:
        raise HTTPException(503, "PySUS não disponível. O backend precisa rodar com Python 3.12 (venv do projeto).")
    try:
        from pysus.ftp.databases.sinan import SINAN as _SINANdb
        _db = _SINANdb().load()
        return sorted(
            [{"codigo": k, "nome": v} for k, v in _db.diseases.items()],
            key=lambda x: x["nome"]
        )
    except Exception as e:
        log.warning(f"Erro ao listar doenças SINAN: {e}")
        return []


@app.get("/api/capacidades")
def get_capacidades():
    """Informa ao frontend quais libs estão disponíveis."""
    return {"pysus_ok": PYSUS_OK, "prophet_ok": PROPHET_OK}


@app.get("/api/ano_limite")
def get_ano_limite():
    """
    Retorna o ano máximo com dados consolidados no DATASUS por sistema.
    O frontend usa esses valores para restringir a seleção de período e
    evitar que o usuário consulte dados preliminares/incompletos.
    """
    defasagens = {"SINAN": 1, "SIM": 1, "SIH": 1, "SINASC": 1, "SIA": 1}
    return {
        sistema: {
            "ano_maximo":  ano,
            "defasagem_anos": defasagens.get(sistema, 2),
            "aviso": (
                f"Dados do {sistema} são consolidados com ~{defasagens.get(sistema, 2)} "
                f"ano(s) de defasagem no DATASUS. "
                f"Selecione anos até {ano} para garantir dados completos."
            ),
        }
        for sistema, ano in ANO_MAXIMO_CONFIAVEL.items()
    }


@app.get("/api/estados")
def get_estados():
    base = list(_EST) if PYSUS_OK else _ESTADOS_LOCAL
    return [{"sigla": s, "nome": n} for s, n in base]


@app.get("/api/cidades/{uf}")
def get_cidades(uf: str):
    """
    Retorna todos os municípios da UF informada (código IBGE de 7 dígitos).
    Fonte primária: API IBGE (5.570 municípios totais).
    Fallback: lista reduzida hardcoded caso a API esteja indisponível.
    Não depende do PySUS.
    """
    lista = _buscar_municipios_ibge(uf.upper())
    return [{"nome": n, "ibge": c} for n, c in lista]


@app.post("/api/download")
def iniciar_download(req: DownloadRequest, background_tasks: BackgroundTasks):
    req_dict = req.model_dump()

    # Cache: se já existe no Supabase, não baixa de novo do DATASUS.
    if req.usar_cache:
        try:
            cached = _try_load_from_supabase(req_dict)
        except Exception as e:
            cached = None
            log.warning(f"  ⚠️  Cache Supabase falhou (vai baixar do DATASUS): {e}")

        if cached and cached.get("meta", {}).get("run_id"):
            job_id = str(cached["meta"]["run_id"])
            jobs[job_id] = {
                "id": job_id,
                "status": "done",
                "progresso": 100,
                "mensagem": "Carregado do Supabase (cache).",
                "resultado": cached,
                "pasta": None,
                "request": req_dict,
            }
            return {"job_id": job_id, "cache": True}

    job_id = str(uuid.uuid4())[:8]
    jobs[job_id] = {
        "id":        job_id,
        "status":    "pending",
        "progresso": 0,
        "mensagem":  "Iniciando...",
        "resultado": None,
        "pasta":     None,
        "request":   req_dict,
    }
    background_tasks.add_task(processar_download, job_id, req_dict)
    return {"job_id": job_id, "cache": False}


@app.get("/api/status/{job_id}")
def get_status(job_id: str):
    if job_id not in jobs:
        raise HTTPException(404, "Job não encontrado")
    j = jobs[job_id]
    return {"id": job_id, "status": j["status"], "progresso": j["progresso"], "mensagem": j["mensagem"]}


@app.get("/api/resultado/{job_id}")
def get_resultado(job_id: str):
    if job_id not in jobs:
        raise HTTPException(404, "Job não encontrado")
    if jobs[job_id]["status"] != "done":
        raise HTTPException(400, "Job ainda não concluído")
    return jobs[job_id]["resultado"]


@app.get("/api/runs")
def list_runs(sistema: str | None = None, limit: int = 200):
    """
    Lista execuções já salvas no Supabase (metadados).
    Usado pela UI (mapa e filtros) para evitar re-download.
    """
    try:
        runs = _list_runs_from_supabase(sistema=sistema, limit=limit)
        return {"ok": True, "runs": runs}
    except Exception as e:
        return {"ok": False, "runs": [], "error": str(e)}


@app.get("/api/export/{job_id}")
def export_xlsx(job_id: str):
    if job_id not in jobs:
        raise HTTPException(404, "Job não encontrado")
    if jobs[job_id]["status"] != "done":
        raise HTTPException(400, "Job ainda não concluído")
    resultado = jobs[job_id].get("resultado")
    if not resultado:
        raise HTTPException(404, "Resultado não encontrado")

    meta = (resultado.get("meta") or {})
    sistema = _slug_filename(str(meta.get("sistema", "SUS")))
    uf = _slug_filename(str(meta.get("uf", "")))
    cidade = _slug_filename(str(meta.get("cidade", "")))
    ano_ini = meta.get("ano_ini", "")
    ano_fim = meta.get("ano_fim", "")
    filename = f"sus_predict_{sistema}_{uf}_{cidade}_{ano_ini}-{ano_fim}_{job_id}.xlsx".strip("_")

    content = _xlsx_bytes(resultado)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.delete("/api/cleanup/{job_id}")
def cleanup(job_id: str):
    if job_id not in jobs:
        raise HTTPException(404, "Job não encontrado")
    pasta = jobs[job_id].get("pasta")
    if pasta:
        shutil.rmtree(pasta, ignore_errors=True)
    del jobs[job_id]
    return {"ok": True, "mensagem": "Dados locais deletados com sucesso."}
